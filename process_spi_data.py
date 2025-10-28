import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from ultralytics import YOLO
from center_paste_dis_img_infer import calculate_pads_distance

import torch
from PIL import Image
import numpy as np
import argparse

# Judgement Thresholds
HIGH_COVER_THRESHOLD = 180
SHORT_DISTANCE_THRESHOLD = 6.6
PAD_CONFIDENCE_THRESHOLD = 0.5
LOW_VOL_OFFSET = -10
HIGH_VOL_OFFSET = 20

def update_based_on_insp_vol(ws_data, ws_to_save, low_vol_offset, high_vol_offset):
    """
    Updates the 'ai_defect_name' column based on 'insp_vol' values in the given worksheet.
    ws_data is a read-only worksheet with formula values.
    ws_to_save is a writeable worksheet to apply changes to.
    """
    orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    header = [cell.value for cell in ws_data[1]]
    try:
        insp_vol_col_idx = header.index('insp_vol') + 1
        ai_defect_name_col_idx = header.index('ai_defect_name') + 1
        vol_l_ng_col_idx = header.index('vol_l_ng') + 1
        vol_h_ng_col_idx = header.index('vol_h_ng') + 1
    except ValueError as e:
        print(f"Error: Missing required column in Excel file: {e}")
        return 0, 0

    high_vol_count = 0
    low_vol_count = 0

    for row_idx in range(2, ws_data.max_row + 1):
        ai_defect_name_cell = ws_to_save.cell(row=row_idx, column=ai_defect_name_col_idx)
        if ai_defect_name_cell.value == 'FM/color':
            continue
        insp_vol_cell = ws_data.cell(row=row_idx, column=insp_vol_col_idx)
        vol_l_ng_cell = ws_data.cell(row=row_idx, column=vol_l_ng_col_idx)
        vol_h_ng_cell = ws_data.cell(row=row_idx, column=vol_h_ng_col_idx)

        if insp_vol_cell.value is not None and vol_l_ng_cell.value is not None and vol_h_ng_cell.value is not None:
            try:
                insp_vol_value = float(insp_vol_cell.value)
                vol_l_ng_value = float(vol_l_ng_cell.value)
                vol_h_ng_value = float(vol_h_ng_cell.value)

                low_vol_threshold = vol_l_ng_value + low_vol_offset
                high_vol_threshold = vol_h_ng_value + high_vol_offset

                if insp_vol_value > high_vol_threshold:
                    ai_defect_name_cell.value = 'high vol'
                    ai_defect_name_cell.fill = orange_fill
                    high_vol_count += 1
                elif insp_vol_value < low_vol_threshold:
                    ai_defect_name_cell.value = 'low vol'
                    ai_defect_name_cell.fill = orange_fill
                    low_vol_count += 1
            except (ValueError, TypeError):
                print(f"Warning: Could not convert insp_vol, vol_l_ng, or vol_h_ng value to float in row {row_idx}. Skipping.")

    return high_vol_count, low_vol_count

def update_based_on_cover_percentage(ws_data, ws_to_save, high_cover_threshold):
    """
    Updates the 'ai_defect_name' column based on 'cover%' values in the given worksheet.
    ws_data is a read-only worksheet with formula values.
    ws_to_save is a writeable worksheet to apply changes to.
    """
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    header = [cell.value for cell in ws_data[1]]
    try:
        cover_col_idx = header.index('cover%') + 1
        ai_defect_name_col_idx = header.index('ai_defect_name') + 1
    except ValueError as e:
        print(f"Error: Missing required column in Excel file: {e}")
        return

    for row_idx in range(2, ws_data.max_row + 1):
        ai_defect_name_cell = ws_to_save.cell(row=row_idx, column=ai_defect_name_col_idx)
        cover_cell = ws_data.cell(row=row_idx, column=cover_col_idx)
        if cover_cell.value is not None:
            try:
                cover_value = float(cover_cell.value)
                if cover_value > high_cover_threshold:
                    ai_defect_name_cell.value = 'high cover'
                    ai_defect_name_cell.fill = yellow_fill
            except (ValueError, TypeError):
                print(f"Warning: Could not convert cover% value '{cover_cell.value}' to float in row {row_idx}. Skipping.")


def update_based_on_distance(ws_data, ws_to_save, center_model, pad_model, root_image_dir, pad_confidence_threshold, high_cover_threshold, short_distance_threshold):
    """
    Updates the 'ai_defect_name' column based on the distance between bounding boxes of detected pads.
    """
    purple_fill = PatternFill(start_color="EAD1DC", end_color="EAD1DC", fill_type="solid")
    header = [cell.value for cell in ws_data[1]]
    
    # Add 'min_pad_distance' column if it doesn't exist
    if 'min_pad_distance' not in header:
        ws_to_save.cell(row=1, column=len(header) + 1, value='min_pad_distance')
        header.append('min_pad_distance') # Update header list for correct indexing
    
    try:
        ai_defect_name_col_idx = header.index('ai_defect_name') + 1
        img_path_col_idx = header.index('img_path') + 1
        cover_col_idx = header.index('cover%') + 1
        min_pad_distance_col_idx = header.index('min_pad_distance') + 1
    except ValueError as e:
        print(f"Error: Missing required column in Excel file: {e}")
        return

    for row_idx in range(2, ws_data.max_row + 1):
        img_path_cell = ws_data.cell(row=row_idx, column=img_path_col_idx)
        if img_path_cell.value:
            full_img_path = os.path.join(root_image_dir, img_path_cell.value)
            if os.path.exists(full_img_path):
                try:
                    # Use the calculate_pads_distance from center_paste_dis_img_infer.py
                    distance = calculate_pads_distance(full_img_path, center_model, pad_model, pad_confidence_threshold)
                
                    # Update min_pad_distance column
                    ws_to_save.cell(row=row_idx, column=min_pad_distance_col_idx, value=distance)

                    cover_cell = ws_data.cell(row=row_idx, column=cover_col_idx)
                    if cover_cell.value is not None:
                        try:
                            cover_value = float(cover_cell.value)
                            if cover_value > high_cover_threshold and distance < short_distance_threshold:
                            # if distance < 7:
                                ai_defect_name_cell = ws_to_save.cell(row=row_idx, column=ai_defect_name_col_idx)
                                ai_defect_name_cell.value = 'short distance'
                                ai_defect_name_cell.fill =  purple_fill
                        except (ValueError, TypeError):
                            print(f"Warning: Could not convert cover% value '{cover_cell.value}' to float in row {row_idx}. Skipping.")
                except Exception as e:
                    print(f"Error calculating pad distance for {img_path_cell.value}: {e}")
                    # Optionally, set min_pad_distance to a default or error value
                    ws_to_save.cell(row=row_idx, column=min_pad_distance_col_idx, value="Error")


def update_ai_juge(ws_data, ws_to_save):
    """
    Updates the 'ai_juge' column to 'NG' if 'ai_defect_name' is not empty.
    """
    header = [cell.value for cell in ws_data[1]]
    
    # Add 'ai_juge' column if it doesn't exist
    if 'ai_juge' not in header:
        ws_to_save.cell(row=1, column=len(header) + 1, value='ai_juge')
        header.append('ai_juge') # Update header list for correct indexing

    try:
        ai_defect_name_col_idx = header.index('ai_defect_name') + 1
        ai_juge_col_idx = header.index('ai_juge') + 1
    except ValueError as e:
        print(f"Error: Missing required column in Excel file: {e}")
        return

    for row_idx in range(2, ws_data.max_row + 1):
        ai_defect_name_cell = ws_to_save.cell(row=row_idx, column=ai_defect_name_col_idx)
        ai_juge_cell = ws_to_save.cell(row=row_idx, column=ai_juge_col_idx)
        
        # Check if ai_defect_name has a value
        if ai_defect_name_cell.value is not None and str(ai_defect_name_cell.value).strip() != "":
            ai_juge_cell.value = "NG"


def Ellipse_area(width, length):
    """Calculates the area of an ellipse given its width and length."""
    return np.pi * width * length / 4

def process_csv_add_area_and_coverage(csv_path):
    """
    Opens a CSV file, adds pad_area and cover% columns, and saves it as an Excel file.
    The output Excel file will have the same name as the CSV file, with the extension changed to .xlsx.
    """
    try:
        df = pd.read_csv(csv_path)

        # 1. Add pad_area column
        if 'Width' in df.columns and 'Length' in df.columns:
            df['pad_area'] = Ellipse_area(df['Width'], df['Length'])
        else:
            print("Error: 'Width' and/or 'Length' columns not found in the CSV file.")
            return

        # 2. Add cover% column
        # The column name 'insp_area_ai(pixel)' is based on the user's request.
        # This may need to be adjusted if the actual column name in the CSV is different.
        pixel_area_col = 'insp_area_ai(pixel)'
        if pixel_area_col in df.columns:
            # The new column should be named 'cover%', as requested.
            df['cover%'] = round(df[pixel_area_col] * 0.8246 * 100 / df['pad_area'], 2)
        else:
            print(f"Error: Column '{pixel_area_col}' not found in the CSV file.")
            return

        # 3. Save it as an Excel file
        excel_path = os.path.splitext(csv_path)[0] + '.xlsx'
        df.to_excel(excel_path, index=False)
        print(f"Successfully processed {csv_path} and saved the output to {excel_path}")

    except FileNotFoundError:
        print(f"Error: The file {csv_path} was not found.")
    except Exception as e:
        print(f"An error occurred: {e}")


def main():
    """
    Main function to process SPI data from a CSV file.
    1. open csv from CSV_FILE_PATH
    2. apply functions to the df, process_csv_add_area_and_coverage, update_based_on_insp_vol, update_based_on_cover_percentage, update_based_on_distance
    3. and save the result to excel with the same name in the same folder.
    """
    parser = argparse.ArgumentParser(description='Process SPI data, adding area, coverage, and distance metrics.')
    parser.add_argument('--csv_path', type=str, required=True, help='Path to the input CSV file.')
    parser.add_argument('--image_root', type=str, required=True, help='Root directory of the images.')
    parser.add_argument('--pad_model_path', type=str, default=r"D:\Dre\SPI_pad_paste\yolo_model\models\gap_model\0714\weights\best.pt", help='Path to the YOLO model for pad detection.')
    parser.add_argument('--center_model_path', type=str, default=r"D:\Dre\SPI_pad_paste\yolo_model\scripts\runs\detect\paste_model\weights\best.pt", help='Path to the YOLO model for center paste detection.')
    parser.add_argument('--low_vol_offset', type=int, default=LOW_VOL_OFFSET, help='Offset for low volume threshold.')
    parser.add_argument('--high_vol_offset', type=int, default=HIGH_VOL_OFFSET, help='Offset for high volume threshold.')
    parser.add_argument('--high_cover_threshold', type=float, default=HIGH_COVER_THRESHOLD, help='Threshold for high cover percentage.')
    parser.add_argument('--short_distance_threshold', type=float, default=SHORT_DISTANCE_THRESHOLD, help='Threshold for short distance between pads.')
    parser.add_argument('--pad_confidence_threshold', type=float, default=PAD_CONFIDENCE_THRESHOLD, help='Confidence threshold for pad detection.')
    args = parser.parse_args()

    # Step 1: Process CSV and save as initial Excel file
    print(f"Processing CSV file: {args.csv_path}")
    process_csv_add_area_and_coverage(args.csv_path)
    
    excel_file_path = os.path.splitext(args.csv_path)[0] + '.xlsx'
    
    # Step 2: Apply further updates to the Excel file
    print(f"Applying updates to Excel file: {excel_file_path}")
    try:
        pad_model = YOLO(args.pad_model_path)
        center_model = YOLO(args.center_model_path)

        wb_data = load_workbook(excel_file_path, data_only=True)
        ws_data = wb_data.active

        wb_to_save = load_workbook(excel_file_path)
        ws_to_save = wb_to_save.active

        print("Updating based on inspection volume...")
        high_vol_count, low_vol_count = update_based_on_insp_vol(ws_data, ws_to_save, args.low_vol_offset, args.high_vol_offset)
        
        print("Updating based on cover percentage...")
        update_based_on_cover_percentage(ws_data, ws_to_save, args.high_cover_threshold)
        
        print("Updating based on distance...")
        update_based_on_distance(ws_data, ws_to_save, center_model, pad_model, args.image_root, args.pad_confidence_threshold, args.high_cover_threshold, args.short_distance_threshold)

        print("Updating AI juge column...")
        update_ai_juge(ws_data, ws_to_save)

        # Step 3: Save the final result
        wb_to_save.save(excel_file_path)
        print(f"Successfully processed and saved updated Excel file: {excel_file_path}")

        total_rows = ws_data.max_row - 1
        
        header = [cell.value for cell in ws_to_save[1]]
        try:
            ai_juge_col_idx = header.index('ai_juge') + 1
            ng_count = 0
            for row_idx in range(2, ws_to_save.max_row + 1):
                if ws_to_save.cell(row=row_idx, column=ai_juge_col_idx).value == 'NG':
                    ng_count += 1
        except ValueError:
            ng_count = 0

        if total_rows > 0:
            rate = (total_rows - (high_vol_count + low_vol_count)) / total_rows
            improve_rate = (total_rows - ng_count) / total_rows
            print(f"High vol count: {high_vol_count}")
            print(f"Low vol count: {low_vol_count}")
            print(f"Pass rate: {rate:.2%}")
            print(f"Improve rate: {improve_rate:.2%}")

    except FileNotFoundError:
        print(f"Error: Excel file not found at {excel_file_path}")
    except Exception as e:
        print(f"An error occurred during Excel update: {e}")

def run_script():
    excel_file_path = r"D:\Dre\SPI_pad_paste\output_results\for_ME\0722\20250523135357_with_images.xlsx"
    pad_model = YOLO(r"D:\Dre\SPI_pad_paste\yolo_model\models\gap_model\0714\weights\best.pt")
    center_model = YOLO(r"D:\Dre\SPI_pad_paste\yolo_model\scripts\runs\detect\paste_model\weights\best.pt")
    image_root = r"D:\Dre\SPI_pad_paste\output_results\for_ME\20250523135357"

    wb_data = load_workbook(excel_file_path, data_only=True)
    ws_data = wb_data.active

    wb_to_save = load_workbook(excel_file_path)
    ws_to_save = wb_to_save.active

    print("Updating based on inspection volume...")
    update_based_on_distance(ws_data, ws_to_save, center_model, pad_model, image_root, PAD_CONFIDENCE_THRESHOLD, HIGH_COVER_THRESHOLD, SHORT_DISTANCE_THRESHOLD)
    wb_to_save.save(excel_file_path)

if __name__ == "__main__":
    main()



